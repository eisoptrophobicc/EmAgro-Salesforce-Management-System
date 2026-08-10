from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.schemas.report import (
    AttendanceReportResponse,
    EmployeeReportResponse,
    ProductivityReportResponse,
)


header_font = Font(bold=True)
status_fills = {
    "Present": PatternFill(
        fill_type="solid",
        fgColor="D9EAD3",
    ),
    "Absent": PatternFill(
        fill_type="solid",
        fgColor="F4CCCC",
    ),
    "Half Day": PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    ),
    "Leave": PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    ),
}
status_pdf_colors = {
    "Present": colors.lightgreen,
    "Absent": colors.lightcoral,
    "Half Day": colors.lightyellow,
    "Leave": colors.lightblue,
}


class ExportService:

    @staticmethod
    def _append_header(sheet, values):
        sheet.append(values)

        for cell in sheet[sheet.max_row]:
            cell.font = header_font

    @staticmethod
    def _append_status_row(sheet, values):
        sheet.append(values)

        fill = status_fills.get(values[0])

        if fill is None:
            return

        for cell in sheet[sheet.max_row]:
            cell.fill = fill

    @staticmethod
    def _append_attendance_breakdown_row(sheet, values, status_start_index):
        sheet.append(values)

        for index, status in enumerate(
            ["Present", "Absent", "Half Day", "Leave"],
            start=status_start_index,
        ):
            sheet.cell(
                row=sheet.max_row,
                column=index,
            ).fill = status_fills[status]

    @staticmethod
    def _autosize_columns(sheet):
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                value = str(cell.value or "")

                if len(value) > max_length:
                    max_length = len(value)

            sheet.column_dimensions[column_letter].width = max_length + 2

    @staticmethod
    def _styled_table(data):
        table = Table(data)
        styles = [
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ]

        for row_index, row in enumerate(data[1:], start=1):
            row_color = status_pdf_colors.get(row[0])

            if row_color is not None:
                styles.append(
                    ("BACKGROUND", (0, row_index), (-1, row_index), row_color)
                )

        for column_index, header in enumerate(data[0]):
            column_color = status_pdf_colors.get(header)

            if column_color is not None:
                styles.append(
                    (
                        "BACKGROUND",
                        (column_index, 1),
                        (column_index, -1),
                        column_color,
                    )
                )

        table.setStyle(TableStyle(styles))
        return table

    @staticmethod
    def productivity_excel(
        report: ProductivityReportResponse,
    ):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Productivity Report"

        sheet["A1"] = "Productivity Report"
        sheet["A1"].font = header_font
        sheet["A2"] = "From"
        sheet["B2"] = report.from_date.isoformat()
        sheet["A3"] = "To"
        sheet["B3"] = report.to_date.isoformat()
        sheet["A5"] = "Attendance"
        sheet["A5"].font = header_font
        sheet.freeze_panes = "A6"
        ExportService._append_header(sheet, ["Status", "Count"])
        ExportService._append_status_row(
            sheet,
            ["Present", report.attendance.present],
        )
        ExportService._append_status_row(
            sheet,
            ["Absent", report.attendance.absent],
        )
        ExportService._append_status_row(
            sheet,
            ["Half Day", report.attendance.half_day],
        )
        ExportService._append_status_row(sheet, ["Leave", report.attendance.leave])
        sheet.append([])
        ExportService._append_header(sheet, ["Task", "Total"])

        for task in report.tasks:
            sheet.append(
                [
                    task.task,
                    task.total,
                ]
            )

        sheet.append([])
        ExportService._append_header(sheet, ["Date", "Total"])

        for item in report.timeline:
            sheet.append(
                [
                    item.date.isoformat(),
                    item.total,
                ]
            )

        buffer = BytesIO()

        ExportService._autosize_columns(sheet)
        workbook.save(buffer)

        buffer.seek(0)

        return buffer

    @staticmethod
    def productivity_pdf(
        report: ProductivityReportResponse,
    ):
        buffer = BytesIO()

        document = SimpleDocTemplate(buffer)

        styles = getSampleStyleSheet()

        elements = []

        elements.append(
            Paragraph(
                "Productivity Report",
                styles["Title"],
            )
        )
        elements.append(Spacer(1, 12))
        elements.append(
            Paragraph(
                f"From: {report.from_date}",
                styles["Normal"],
            )
        )
        elements.append(
            Paragraph(
                f"To: {report.to_date}",
                styles["Normal"],
            )
        )
        elements.append(Spacer(1, 20))
        attendance_data = [
            ["Status", "Count"],
            ["Present", report.attendance.present],
            ["Absent", report.attendance.absent],
            ["Half Day", report.attendance.half_day],
            ["Leave", report.attendance.leave],
        ]
        elements.append(ExportService._styled_table(attendance_data))
        elements.append(Spacer(1, 20))
        task_data = [
            ["Task", "Total"],
        ]

        for task in report.tasks:
            task_data.append(
                [
                    task.task,
                    task.total,
                ]
            )

        elements.append(ExportService._styled_table(task_data))
        elements.append(Spacer(1, 20))

        document.build(elements)

        buffer.seek(0)

        return buffer

    @staticmethod
    def attendance_excel(
        report: AttendanceReportResponse,
    ):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Attendance Report"

        sheet["A1"] = "Attendance Report"
        sheet["A1"].font = header_font
        sheet["A2"] = "From"
        sheet["B2"] = report.from_date.isoformat()
        sheet["A3"] = "To"
        sheet["B3"] = report.to_date.isoformat()
        sheet["A5"] = "Attendance Summary"
        sheet["A5"].font = header_font
        sheet.freeze_panes = "A6"
        ExportService._append_header(sheet, ["Status", "Count"])
        ExportService._append_status_row(
            sheet,
            ["Present", report.summary.present],
        )
        ExportService._append_status_row(
            sheet,
            ["Absent", report.summary.absent],
        )
        ExportService._append_status_row(
            sheet,
            ["Half Day", report.summary.half_day],
        )
        ExportService._append_status_row(sheet, ["Leave", report.summary.leave])
        sheet.append([])
        ExportService._append_header(
            sheet,
            [
                "Employee",
                "Present",
                "Absent",
                "Half Day",
                "Leave",
            ]
        )

        for employee in report.employees:
            ExportService._append_attendance_breakdown_row(
                sheet,
                [
                    employee.employee,
                    employee.present,
                    employee.absent,
                    employee.half_day,
                    employee.leave,
                ],
                2,
            )

        buffer = BytesIO()

        ExportService._autosize_columns(sheet)
        workbook.save(buffer)

        buffer.seek(0)

        return buffer

    @staticmethod
    def attendance_pdf(
        report: AttendanceReportResponse,
    ):
        buffer = BytesIO()

        document = SimpleDocTemplate(buffer)

        styles = getSampleStyleSheet()

        elements = []
        elements.append(
            Paragraph(
                "Attendance Report",
                styles["Title"],
            )
        )
        elements.append(Spacer(1, 12))
        elements.append(
            Paragraph(
                f"From: {report.from_date}",
                styles["Normal"],
            )
        )
        elements.append(
            Paragraph(
                f"To: {report.to_date}",
                styles["Normal"],
            )
        )
        elements.append(Spacer(1, 20))
        attendance_data = [
            ["Status", "Count"],
            ["Present", report.summary.present],
            ["Absent", report.summary.absent],
            ["Half Day", report.summary.half_day],
            ["Leave", report.summary.leave],
        ]
        elements.append(ExportService._styled_table(attendance_data))
        elements.append(Spacer(1, 20))
        employee_data = [
            [
                "Employee",
                "Present",
                "Absent",
                "Half Day",
                "Leave",
            ]
        ]

        for employee in report.employees:
            employee_data.append(
                [
                    employee.employee,
                    employee.present,
                    employee.absent,
                    employee.half_day,
                    employee.leave,
                ]
            )

        elements.append(ExportService._styled_table(employee_data))

        document.build(elements)

        buffer.seek(0)

        return buffer

    @staticmethod
    def employee_excel(
        report: EmployeeReportResponse,
    ):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Employee Report"

        sheet["A1"] = "Employee Report"
        sheet["A1"].font = header_font
        sheet["A2"] = "From"
        sheet["B2"] = report.from_date.isoformat()
        sheet["A3"] = "To"
        sheet["B3"] = report.to_date.isoformat()
        sheet.freeze_panes = "A5"
        sheet.append([])
        ExportService._append_header(
            sheet,
            [
                "Employee Code",
                "Employee",
                "Designation",
                "Present",
                "Absent",
                "Half Day",
                "Leave",
            ]
        )

        for employee in report.employees:
            ExportService._append_attendance_breakdown_row(
                sheet,
                [
                    employee.employee_code,
                    employee.full_name,
                    employee.designation,
                    employee.present,
                    employee.absent,
                    employee.half_day,
                    employee.leave,
                ],
                4,
            )
            ExportService._append_header(sheet, ["Task", "Total"])

            for task in employee.tasks:
                sheet.append(
                    [
                        task.task,
                        task.total,
                    ]
                )

            sheet.append([])

        buffer = BytesIO()

        ExportService._autosize_columns(sheet)
        workbook.save(buffer)

        buffer.seek(0)

        return buffer

    @staticmethod
    def employee_pdf(
        report: EmployeeReportResponse,
    ):
        buffer = BytesIO()

        document = SimpleDocTemplate(buffer)

        styles = getSampleStyleSheet()

        elements = []
        elements.append(
            Paragraph(
                "Employee Report",
                styles["Title"],
            )
        )
        elements.append(Spacer(1, 12))
        elements.append(
            Paragraph(
                f"From: {report.from_date}",
                styles["Normal"],
            )
        )
        elements.append(
            Paragraph(
                f"To: {report.to_date}",
                styles["Normal"],
            )
        )
        elements.append(Spacer(1, 20))

        for employee in report.employees:
            elements.append(
                Paragraph(
                    f"{employee.full_name} ({employee.employee_code})",
                    styles["Heading2"],
                )
            )
            elements.append(
                Paragraph(
                    employee.designation,
                    styles["Normal"],
                )
            )
            attendance_data = [
                ["Status", "Count"],
                ["Present", employee.present],
                ["Absent", employee.absent],
                ["Half Day", employee.half_day],
                ["Leave", employee.leave],
            ]
            elements.append(ExportService._styled_table(attendance_data))
            elements.append(Spacer(1, 10))
            task_data = [["Task", "Total"]]

            for task in employee.tasks:
                task_data.append(
                    [
                        task.task,
                        task.total,
                    ]
                )

            elements.append(ExportService._styled_table(task_data))
            elements.append(Spacer(1, 20))

        document.build(elements)

        buffer.seek(0)

        return buffer
